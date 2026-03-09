
from datetime import date
from io import BytesIO
from pathlib import Path
import textwrap

import matplotlib.pyplot as plt
import pandas as pd
import streamlit as st
from matplotlib.backends.backend_pdf import PdfPages


st.set_page_config(
    page_title="Project Cost Intelligence System",
    page_icon="🏗️",
    layout="wide",
)

PROJECT_TYPE_SPLITS = {
    "General Contracting": {
        "Materials": 40,
        "Labor": 25,
        "Transportation": 5,
        "Office Expense": 5,
        "Salaries / Overheads": 10,
        "Company Profit": 10,
        "Contingency": 5,
    },
    "Pipeline Project": {
        "Materials": 50,
        "Labor": 20,
        "Transportation": 10,
        "Office Expense": 5,
        "Salaries / Overheads": 5,
        "Company Profit": 7,
        "Contingency": 3,
    },
    "Civil Construction": {
        "Materials": 45,
        "Labor": 30,
        "Transportation": 5,
        "Office Expense": 5,
        "Salaries / Overheads": 8,
        "Company Profit": 5,
        "Contingency": 2,
    },
    "Mechanical Installation": {
        "Materials": 38,
        "Labor": 32,
        "Transportation": 6,
        "Office Expense": 5,
        "Salaries / Overheads": 9,
        "Company Profit": 7,
        "Contingency": 3,
    },
    "Maintenance Contract": {
        "Materials": 25,
        "Labor": 40,
        "Transportation": 8,
        "Office Expense": 7,
        "Salaries / Overheads": 10,
        "Company Profit": 7,
        "Contingency": 3,
    },
    "EPC Project": {
        "Materials": 42,
        "Labor": 22,
        "Transportation": 6,
        "Office Expense": 6,
        "Salaries / Overheads": 12,
        "Company Profit": 8,
        "Contingency": 4,
    },
}

CURRENCY_SYMBOLS = {
    "AED": "AED",
    "USD": "$",
    "INR": "₹",
}


def get_allocation_status(total_pct: float) -> str:
    if abs(total_pct - 100) < 1e-9:
        return "Balanced"
    if total_pct < 100:
        return "Under Allocated"
    return "Over Allocated"


def build_budget_dataframe(total_budget: float, allocations: dict) -> pd.DataFrame:
    rows = []
    for category, pct in allocations.items():
        rows.append(
            {
                "Category": category,
                "Allocation %": round(float(pct), 2),
                "Amount": round(total_budget * float(pct) / 100, 2),
            }
        )
    return pd.DataFrame(rows)


def analyze_budget(allocations: dict, total_pct: float) -> list[str]:
    insights = []
    status = get_allocation_status(total_pct)

    if status == "Balanced":
        insights.append("✅ Total allocation is perfectly balanced at 100%.")
    elif status == "Under Allocated":
        insights.append("⚠️ Total allocation is below 100%. Some budget is still unassigned.")
    else:
        insights.append("❌ Total allocation exceeds 100%. Adjust the percentages to match the budget.")

    profit = allocations["Company Profit"]
    labor = allocations["Labor"]
    materials = allocations["Materials"]
    contingency = allocations["Contingency"]
    transport = allocations["Transportation"]

    if profit < 8:
        insights.append("⚠️ Profit margin is below 8%. This may be too tight for comfortable execution.")
    elif profit <= 15:
        insights.append("✅ Profit margin looks healthy for a practical contracting estimate.")
    else:
        insights.append("⚠️ Profit margin is quite high. Double-check competitiveness and client expectations.")

    if labor < 20:
        insights.append("⚠️ Labor allocation is low. Execution quality or manpower coverage may suffer.")
    elif labor > 35:
        insights.append("⚠️ Labor allocation is very high. Verify productivity assumptions and manpower planning.")

    if materials < 30:
        insights.append("⚠️ Material allocation is low. Review BOQ realism and procurement assumptions.")
    elif materials > 60:
        insights.append("⚠️ Material allocation is very high. Check whether supply cost is dominating the project.")

    if contingency < 5:
        insights.append("⚠️ Contingency is low. This leaves little room for site surprises or change orders.")
    else:
        insights.append("✅ Contingency reserve provides a reasonable safety buffer.")

    if transport > 10:
        insights.append("⚠️ Transportation cost is high. Recheck logistics, fuel, and delivery assumptions.")

    return insights


def generate_report_text(
    project_name: str,
    project_type: str,
    company_name: str,
    client_name: str,
    project_reference: str,
    currency_symbol: str,
    total_budget: float,
    allocations: dict,
    total_pct: float,
) -> dict[str, str]:
    profit_pct = allocations["Company Profit"]
    profit_amount = total_budget * profit_pct / 100
    allocation_status = get_allocation_status(total_pct).lower()

    estimate_summary = (
        f"Project Reference: {project_reference}. Client: {client_name}. "
        f"Based on the submitted project details for {project_name}, categorized under {project_type}, "
        f"{company_name} has prepared a preliminary budget estimate with a total projected value of "
        f"{currency_symbol} {total_budget:,.2f}. The estimate has been distributed across key execution heads "
        f"including materials, labor, transportation, office expenses, overheads, contingency, and company profit. "
        f"The current model reflects an expected profit margin of {profit_pct:.2f}% with an estimated profit value "
        f"of {currency_symbol} {profit_amount:,.2f}. Overall allocation status is presently {allocation_status}, "
        f"subject to final technical and commercial review."
    )

    proposal_note = (
        f"Reference {project_reference} has been prepared for {client_name}. "
        f"We are pleased to submit this preliminary commercial budget consideration for {project_name}. "
        f"This estimate has been developed based on the selected project type, expected execution requirements, "
        f"material demand, labor deployment, logistics considerations, and administrative overheads. "
        f"The proposed allocation is intended to maintain practical project execution while preserving commercial "
        f"viability for {company_name}. Any final quotation or commercial submission should be validated against "
        f"approved scope, drawings, specifications, and prevailing market rates before issue."
    )

    disclaimer_note = (
        "This budget estimate is prepared solely for planning, review, and preliminary commercial evaluation purposes. "
        "Final values may vary depending on site conditions, actual quantities, client requirements, specification "
        "changes, transportation conditions, authority approvals, and material price fluctuations at the time of execution. "
        "Any variation in scope, timeline, procurement conditions, or execution methodology may lead to revision "
        "of the final commercial offer and agreement terms."
    )

    internal_note = (
        f"Internal review reference: {project_reference}. "
        f"From an internal review standpoint, the current allocation for {project_name} appears "
        f"{allocation_status} with a projected profit margin of {profit_pct:.2f}%. "
        "Management is advised to review category distribution, execution risks, and contingency adequacy "
        "before final approval or client submission."
    )

    return {
        "Estimate Summary": estimate_summary,
        "Commercial Proposal Note": proposal_note,
        "Terms / Disclaimer Note": disclaimer_note,
        "Internal Approval Note": internal_note,
    }


def create_pdf_report(
    project_name: str,
    project_type: str,
    company_name: str,
    client_name: str,
    project_reference: str,
    currency_code: str,
    currency_symbol: str,
    author_name: str,
    today: str,
    total_budget: float,
    allocations: dict,
    df: pd.DataFrame,
    insights: list[str],
    report_sections: dict[str, str],
) -> BytesIO:
    pdf_buffer = BytesIO()
    profit_amount = total_budget * allocations["Company Profit"] / 100
    total_pct = sum(allocations.values())
    allocation_status = get_allocation_status(total_pct)

    with PdfPages(pdf_buffer) as pdf:
        fig1 = plt.figure(figsize=(8.27, 11.69))
        ax1 = fig1.add_axes([0, 0, 1, 1])
        ax1.axis("off")

        y = 0.96
        ax1.text(0.5, y, company_name, ha="center", va="top", fontsize=20, fontweight="bold")
        y -= 0.04
        ax1.text(0.5, y, "PROJECT COST INTELLIGENCE REPORT", ha="center", va="top", fontsize=14, fontweight="bold")
        y -= 0.06

        left_lines = [
            f"Project Name: {project_name}",
            f"Client Name: {client_name}",
            f"Project Reference: {project_reference}",
            f"Project Type: {project_type}",
            f"Date: {today}",
            f"Prepared By: {author_name}",
            f"Currency: {currency_code}",
        ]
        right_lines = [
            f"Total Budget: {currency_symbol} {total_budget:,.2f}",
            f"Profit Margin: {allocations['Company Profit']:.2f}%",
            f"Profit Amount: {currency_symbol} {profit_amount:,.2f}",
            f"Allocation Status: {allocation_status}",
        ]

        y_left = y
        for line in left_lines:
            ax1.text(0.08, y_left, line, fontsize=10, va="top")
            y_left -= 0.03

        y_right = y
        for line in right_lines:
            ax1.text(0.58, y_right, line, fontsize=10, va="top", fontweight="bold")
            y_right -= 0.04

        y = min(y_left, y_right) - 0.03
        ax1.text(0.08, y, "Budget Intelligence", fontsize=12, fontweight="bold", va="top")
        y -= 0.03

        for insight in insights:
            wrapped = textwrap.fill(insight, width=90)
            ax1.text(0.09, y, wrapped, fontsize=9, va="top")
            y -= 0.04 + 0.012 * wrapped.count("\n")

        pdf.savefig(fig1, bbox_inches="tight")
        plt.close(fig1)

        fig2 = plt.figure(figsize=(8.27, 11.69))
        ax2 = fig2.add_axes([0.05, 0.05, 0.9, 0.9])
        ax2.axis("off")
        ax2.set_title("Detailed Budget Breakdown", fontsize=14, fontweight="bold", pad=20)

        display_df = df.copy()
        display_df["Allocation %"] = display_df["Allocation %"].map(lambda x: f"{x:.2f}%")
        display_df["Amount"] = display_df["Amount"].map(lambda x: f"{currency_symbol} {x:,.2f}")

        table = ax2.table(
            cellText=display_df.values,
            colLabels=display_df.columns,
            cellLoc="center",
            colLoc="center",
            loc="upper center",
        )
        table.auto_set_font_size(False)
        table.set_fontsize(9)
        table.scale(1, 1.6)

        for (row, _col), cell in table.get_celld().items():
            if row == 0:
                cell.set_text_props(weight="bold")
                cell.set_facecolor("#D9EAF7")
            elif row > 0:
                category = display_df.iloc[row - 1, 0]
                if category == "Company Profit":
                    cell.set_facecolor("#FFF2CC")
                elif category == "Contingency":
                    cell.set_facecolor("#E2F0D9")

        pdf.savefig(fig2, bbox_inches="tight")
        plt.close(fig2)

        fig3 = plt.figure(figsize=(8.27, 11.69))
        ax3 = fig3.add_axes([0.08, 0.53, 0.84, 0.36])
        ax3.pie(df["Amount"], labels=df["Category"], autopct="%1.1f%%", startangle=90)
        ax3.set_title("Budget Distribution")

        ax4 = fig3.add_axes([0.08, 0.09, 0.84, 0.32])
        ax4.bar(df["Category"], df["Amount"])
        ax4.set_title(f"Category-wise Budget Amount ({currency_code})")
        ax4.set_ylabel(currency_code)
        ax4.tick_params(axis="x", rotation=45)

        pdf.savefig(fig3, bbox_inches="tight")
        plt.close(fig3)

        fig4 = plt.figure(figsize=(8.27, 11.69))
        ax5 = fig4.add_axes([0, 0, 1, 1])
        ax5.axis("off")
        ax5.text(0.5, 0.96, "AUTO-GENERATED REPORT DRAFT", ha="center", va="top", fontsize=14, fontweight="bold")

        y = 0.9
        for title, content in report_sections.items():
            ax5.text(0.07, y, title, fontsize=11, fontweight="bold", va="top")
            y -= 0.03
            wrapped = textwrap.fill(content, width=100)
            ax5.text(0.08, y, wrapped, fontsize=9, va="top")
            y -= 0.08 + 0.014 * wrapped.count("\n")

        pdf.savefig(fig4, bbox_inches="tight")
        plt.close(fig4)

    pdf_buffer.seek(0)
    return pdf_buffer


def create_excel_report(
    project_name: str,
    project_type: str,
    company_name: str,
    client_name: str,
    project_reference: str,
    currency_code: str,
    currency_symbol: str,
    author_name: str,
    today: str,
    total_budget: float,
    allocations: dict,
    df: pd.DataFrame,
    insights: list[str],
    report_sections: dict[str, str],
    logo_path: str | None = None,
) -> BytesIO:
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    output = BytesIO()
    profit_pct = allocations["Company Profit"]
    profit_amount = total_budget * profit_pct / 100
    total_allocation_pct = round(sum(allocations.values()), 2)
    allocation_status = get_allocation_status(total_allocation_pct)
    insights_with_status = [f"Allocation Status: {allocation_status}"] + insights

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Budget Breakdown", startrow=7)
        pd.DataFrame({"Budget Insights": insights_with_status}).to_excel(
            writer, index=False, sheet_name="Budget Insights", startrow=3
        )
        wb = writer.book

        dark_fill = PatternFill("solid", fgColor="1F4E78")
        mid_fill = PatternFill("solid", fgColor="D9EAF7")
        soft_fill = PatternFill("solid", fgColor="EDF4FB")
        gold_fill = PatternFill("solid", fgColor="FFF2CC")
        green_fill = PatternFill("solid", fgColor="E2F0D9")
        red_fill = PatternFill("solid", fgColor="FDE9E7")
        gray_fill = PatternFill("solid", fgColor="F3F3F3")

        title_font = Font(color="FFFFFF", bold=True, size=16)
        section_font = Font(bold=True, size=11)
        header_font = Font(bold=True)
        big_font = Font(bold=True, size=13)

        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

        thin = Side(style="thin", color="B7B7B7")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws = wb.create_sheet("Executive Summary", 0)
        ws.sheet_view.showGridLines = False
        for col, width in {"A": 16, "B": 24, "C": 24, "D": 24, "E": 24, "F": 24}.items():
            ws.column_dimensions[col].width = width

        ws["A1"].fill = dark_fill
        ws["A1"].border = border
        ws["A1"].alignment = center
        ws.merge_cells("B1:F1")
        ws["B1"] = f"{company_name} - PROJECT COST INTELLIGENCE REPORT"
        ws["B1"].fill = dark_fill
        ws["B1"].font = title_font
        ws["B1"].alignment = center
        ws.row_dimensions[1].height = 34

        if logo_path and Path(logo_path).exists():
            try:
                logo = XLImage(logo_path)
                logo.width = 72
                logo.height = 72
                ws.add_image(logo, "A1")
            except Exception:
                ws["A1"] = "LOGO"
                ws["A1"].font = header_font
                ws["A1"].alignment = center
        else:
            ws["A1"] = "LOGO"
            ws["A1"].font = header_font
            ws["A1"].alignment = center

        ws.merge_cells("A3:C3")
        ws["A3"] = "PROJECT INFORMATION"
        ws["A3"].fill = mid_fill
        ws["A3"].font = section_font
        ws["A3"].alignment = center

        ws.merge_cells("D3:F3")
        ws["D3"] = "KEY FINANCIAL METRICS"
        ws["D3"].fill = gold_fill
        ws["D3"].font = section_font
        ws["D3"].alignment = center

        info_rows = [
            ("Project Name", project_name),
            ("Project Reference", project_reference),
            ("Client Name", client_name),
            ("Project Type", project_type),
            ("Company Name", company_name),
            ("Author", author_name),
            ("Date", today),
            ("Currency", currency_code),
        ]
        metric_rows = [
            ("Total Budget", total_budget),
            ("Profit %", profit_pct / 100),
            ("Profit Amount", profit_amount),
            ("Allocation %", total_allocation_pct / 100),
            ("Status", allocation_status),
        ]

        start_row = 5
        for i, (label, value) in enumerate(info_rows, start=start_row):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            ws[f"A{i}"].fill = soft_fill
            ws[f"A{i}"].font = header_font
            ws[f"A{i}"].border = border
            ws[f"B{i}"].border = border
            ws[f"C{i}"].border = border
            ws[f"A{i}"].alignment = left
            ws[f"B{i}"].alignment = left

        for i, (label, value) in enumerate(metric_rows, start=start_row):
            ws[f"D{i}"] = label
            ws[f"E{i}"] = value
            ws[f"F{i}"] = ""
            ws[f"D{i}"].fill = gray_fill
            ws[f"D{i}"].font = header_font
            ws[f"D{i}"].border = border
            ws[f"E{i}"].border = border
            ws[f"F{i}"].border = border
            ws[f"D{i}"].alignment = left
            ws[f"E{i}"].alignment = center if label != "Status" else left

            if label in {"Total Budget", "Profit Amount"}:
                ws[f"E{i}"].number_format = f'"{currency_symbol}" #,##0.00'
                ws[f"E{i}"].font = big_font
            elif label in {"Profit %", "Allocation %"}:
                ws[f"E{i}"].number_format = "0.00%"
                ws[f"E{i}"].font = big_font
            elif label == "Status":
                ws[f"E{i}"].fill = green_fill if allocation_status == "Balanced" else red_fill
                ws[f"E{i}"].font = header_font

        note_header_row = start_row + max(len(info_rows), len(metric_rows)) + 2
        note_body_row = note_header_row + 1
        ws.merge_cells(f"A{note_header_row}:F{note_header_row}")
        ws[f"A{note_header_row}"] = "EXECUTIVE NOTE"
        ws[f"A{note_header_row}"].fill = mid_fill
        ws[f"A{note_header_row}"].font = section_font
        ws[f"A{note_header_row}"].alignment = center

        note_text = (
            f"This report summarizes the recommended allocation for a {project_type.lower()} under {company_name} "
            f"for client {client_name}. Current allocation status is {allocation_status.lower()}, with an estimated "
            f"profit of {profit_pct:.2f}% and profit amount of {currency_symbol} {profit_amount:,.2f}."
        )
        ws.merge_cells(f"A{note_body_row}:F{note_body_row + 2}")
        ws[f"A{note_body_row}"] = note_text
        ws[f"A{note_body_row}"].alignment = wrap_left
        ws[f"A{note_body_row}"].border = border

        bd = wb["Budget Breakdown"]
        bd.sheet_view.showGridLines = False
        for col, width in {"A": 30, "B": 16, "C": 22, "D": 18}.items():
            bd.column_dimensions[col].width = width

        bd.merge_cells("A1:D1")
        bd["A1"] = "DETAILED BUDGET BREAKDOWN"
        bd["A1"].fill = dark_fill
        bd["A1"].font = title_font
        bd["A1"].alignment = center

        bd["A3"] = "Project Name"
        bd["B3"] = project_name
        bd["A4"] = "Project Reference"
        bd["B4"] = project_reference
        bd["A5"] = "Client Name"
        bd["B5"] = client_name
        bd["A6"] = "Project Type"
        bd["B6"] = project_type
        bd["D3"] = "Prepared By"
        bd["D4"] = author_name
        bd["D5"] = "Prepared On"
        bd["D6"] = today

        for cell in ["A3", "A4", "A5", "A6", "D3", "D5"]:
            bd[cell].fill = soft_fill
            bd[cell].font = header_font
        for cell in ["A3", "B3", "A4", "B4", "A5", "B5", "A6", "B6", "D3", "D4", "D5", "D6"]:
            bd[cell].border = border

        table_header_row = 8
        bd["D8"] = "Remarks"
        for cell in bd[table_header_row]:
            cell.fill = mid_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for row in range(table_header_row + 1, table_header_row + 1 + len(df)):
            bd[f"B{row}"].number_format = "0.00"
            bd[f"C{row}"].number_format = f'"{currency_symbol}" #,##0.00'
            bd[f"D{row}"] = "Reviewed"
            for col in "ABCD":
                bd[f"{col}{row}"].border = border

            category = bd[f"A{row}"].value
            if category == "Company Profit":
                for col in "ABCD":
                    bd[f"{col}{row}"].fill = gold_fill
                    bd[f"{col}{row}"].font = header_font
            elif category == "Contingency":
                for col in "ABCD":
                    bd[f"{col}{row}"].fill = green_fill

        total_row = table_header_row + 1 + len(df)
        bd[f"A{total_row}"] = "Total Allocation"
        bd[f"B{total_row}"] = total_allocation_pct
        bd[f"C{total_row}"] = df["Amount"].sum()
        bd[f"D{total_row}"] = allocation_status
        for col in "ABCD":
            bd[f"{col}{total_row}"].fill = gray_fill
            bd[f"{col}{total_row}"].font = header_font
            bd[f"{col}{total_row}"].border = border

        bd[f"B{total_row}"].number_format = "0.00"
        bd[f"C{total_row}"].number_format = f'"{currency_symbol}" #,##0.00'
        bd.freeze_panes = "A8"

        ins = wb["Budget Insights"]
        ins.sheet_view.showGridLines = False
        ins.column_dimensions["A"].width = 110
        ins["A1"] = "BUDGET INTELLIGENCE & RISK NOTES"
        ins["A1"].fill = dark_fill
        ins["A1"].font = title_font
        ins["A1"].alignment = center
        ins["A4"].fill = mid_fill
        ins["A4"].font = header_font
        ins["A4"].border = border

        for row in range(5, 5 + len(insights_with_status)):
            ins[f"A{row}"].border = border
            ins[f"A{row}"].alignment = wrap_left
            value = str(ins[f"A{row}"].value)
            if "✅" in value or "Balanced" in value:
                ins[f"A{row}"].fill = green_fill
            elif "⚠️" in value or "❌" in value or "Over" in value or "Under" in value:
                ins[f"A{row}"].fill = red_fill

        ins.freeze_panes = "A4"

        rpt = wb.create_sheet("Report Draft")
        rpt.sheet_view.showGridLines = False
        rpt.column_dimensions["A"].width = 28
        rpt.column_dimensions["B"].width = 110
        rpt.merge_cells("A1:B1")
        rpt["A1"] = "AUTO-GENERATED REPORT DRAFT"
        rpt["A1"].fill = dark_fill
        rpt["A1"].font = title_font
        rpt["A1"].alignment = center

        report_row = 3
        for title, content in report_sections.items():
            rpt[f"A{report_row}"] = title
            rpt[f"A{report_row}"].fill = mid_fill
            rpt[f"A{report_row}"].font = header_font
            rpt[f"A{report_row}"].border = border
            rpt[f"B{report_row}"] = content
            rpt[f"B{report_row}"].alignment = wrap_left
            rpt[f"B{report_row}"].border = border
            rpt.row_dimensions[report_row].height = 54
            report_row += 2

        ch = wb.create_sheet("Charts")
        ch.sheet_view.showGridLines = False
        ch.merge_cells("A1:H1")
        ch["A1"] = "BUDGET VISUALS"
        ch["A1"].fill = dark_fill
        ch["A1"].font = title_font
        ch["A1"].alignment = center
        ch["A2"] = "Category"
        ch["B2"] = "Amount"
        ch["C2"] = "Allocation %"

        for cell in ch[2]:
            cell.fill = mid_fill
            cell.font = header_font
            cell.border = border

        for idx, (_, row) in enumerate(df.iterrows(), start=3):
            ch[f"A{idx}"] = row["Category"]
            ch[f"B{idx}"] = row["Amount"]
            ch[f"C{idx}"] = row["Allocation %"] / 100
            ch[f"B{idx}"].number_format = f'"{currency_symbol}" #,##0.00'
            ch[f"C{idx}"].number_format = "0.00%"

        pie = PieChart()
        pie_labels = Reference(ch, min_col=1, min_row=3, max_row=2 + len(df))
        pie_data = Reference(ch, min_col=2, min_row=2, max_row=2 + len(df))
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(pie_labels)
        pie.title = "Budget Distribution"
        pie.height = 9
        pie.width = 12
        ch.add_chart(pie, "E3")

        bar = BarChart()
        bar.type = "col"
        bar.style = 10
        bar.title = "Allocation by Category"
        bar.y_axis.title = f"Amount ({currency_code})"
        bar.x_axis.title = "Category"
        bar_data = Reference(ch, min_col=2, min_row=2, max_row=2 + len(df))
        bar_labels = Reference(ch, min_col=1, min_row=3, max_row=2 + len(df))
        bar.add_data(bar_data, titles_from_data=True)
        bar.set_categories(bar_labels)
        bar.height = 9
        bar.width = 14
        ch.add_chart(bar, "E20")

    output.seek(0)
    return output


st.title("🏗️ Project Cost Intelligence System")
st.caption("Smart budget allocation, profit analysis, and planning dashboard for engineering / contracting projects.")

st.subheader("Project Details")
col1, col2 = st.columns(2)

with col1:
    project_name = st.text_input("Project Name", value="New Engineering Project")
    client_name = st.text_input("Client Name", value="Client Name")
    company_name = st.text_input("Company Name", value="TKM")
    logo_path_input = st.text_input("Company Logo Path (optional)", value="")

with col2:
    author_name = "Abdulla Zahin"
    today = date.today().strftime("%d %B %Y")
    project_reference = st.text_input("Project Reference", value="TKM-EST-2026-001")
    currency_code = st.selectbox("Currency", options=["AED", "USD", "INR"], index=0)
    st.text_input("Author", value=author_name, disabled=True)
    st.text_input("Date", value=today, disabled=True)

currency_symbol = CURRENCY_SYMBOLS[currency_code]

st.subheader("Budget Input")
project_type = st.selectbox("Project Type", options=list(PROJECT_TYPE_SPLITS.keys()), index=0)

if "last_project_type" not in st.session_state:
    st.session_state.last_project_type = project_type
if "allocations_state" not in st.session_state:
    st.session_state.allocations_state = PROJECT_TYPE_SPLITS[project_type].copy()
if project_type != st.session_state.last_project_type:
    st.session_state.allocations_state = PROJECT_TYPE_SPLITS[project_type].copy()
    st.session_state.last_project_type = project_type

selected_split = st.session_state.allocations_state

total_budget = st.number_input(
    "Total Project Budget",
    min_value=0.0,
    value=500000.0,
    step=10000.0,
    format="%.2f",
)

st.markdown("### Recommended Allocation")
st.write(f"Smart recommendation is loaded for **{project_type}**. You can still adjust the percentages if needed.")

alloc_col1, alloc_col2 = st.columns(2)
allocations = {}
items = list(selected_split.items())
mid = len(items) // 2 + len(items) % 2

with alloc_col1:
    for category, _ in items[:mid]:
        widget_key = f"alloc_{category}"
        allocations[category] = st.number_input(
            f"{category} (%)",
            min_value=0.0,
            max_value=100.0,
            value=float(st.session_state.allocations_state[category]),
            step=1.0,
            key=widget_key,
        )
        st.session_state.allocations_state[category] = allocations[category]

with alloc_col2:
    for category, _ in items[mid:]:
        widget_key = f"alloc_{category}"
        allocations[category] = st.number_input(
            f"{category} (%)",
            min_value=0.0,
            max_value=100.0,
            value=float(st.session_state.allocations_state[category]),
            step=1.0,
            key=widget_key,
        )
        st.session_state.allocations_state[category] = allocations[category]

total_pct = sum(allocations.values())
remaining_pct = 100 - total_pct

df = build_budget_dataframe(total_budget, allocations)
insights = analyze_budget(allocations, total_pct)
report_sections = generate_report_text(
    project_name=project_name,
    project_type=project_type,
    company_name=company_name,
    client_name=client_name,
    project_reference=project_reference,
    currency_symbol=currency_symbol,
    total_budget=total_budget,
    allocations=allocations,
    total_pct=total_pct,
)

info1, info2, info3, info4 = st.columns(4)
info1.metric("Total Allocation %", f"{total_pct:.2f}%")
info2.metric("Unallocated / Excess %", f"{remaining_pct:.2f}%")
info3.metric("Estimated Profit %", f"{allocations['Company Profit']:.2f}%")
info4.metric("Estimated Profit Amount", f"{currency_symbol} {total_budget * allocations['Company Profit'] / 100:,.2f}")

st.subheader("Budget Breakdown")
st.dataframe(df, use_container_width=True, hide_index=True)

st.subheader("Budget Intelligence")
for item in insights:
    st.write(item)

st.subheader("Project Summary")
summary1, summary2, summary3, summary4 = st.columns(4)
profit_amount = total_budget * allocations["Company Profit"] / 100
materials_amount = total_budget * allocations["Materials"] / 100
labor_amount = total_budget * allocations["Labor"] / 100
contingency_amount = total_budget * allocations["Contingency"] / 100

summary1.metric("Total Budget", f"{currency_symbol} {total_budget:,.2f}")
summary2.metric("Materials", f"{currency_symbol} {materials_amount:,.2f}")
summary3.metric("Labor", f"{currency_symbol} {labor_amount:,.2f}")
summary4.metric("Contingency", f"{currency_symbol} {contingency_amount:,.2f}")
st.metric("Company Profit", f"{currency_symbol} {profit_amount:,.2f}", f"{allocations['Company Profit']:.2f}% margin")

st.subheader("Visual Dashboard")
chart_col1, chart_col2 = st.columns(2)

with chart_col1:
    fig1, ax1 = plt.subplots(figsize=(6, 6))
    ax1.pie(df["Amount"], labels=df["Category"], autopct="%1.1f%%", startangle=90)
    ax1.set_title("Budget Distribution")
    st.pyplot(fig1)
    plt.close(fig1)

with chart_col2:
    fig2, ax2 = plt.subplots(figsize=(8, 5))
    ax2.bar(df["Category"], df["Amount"])
    ax2.set_title("Category-wise Budget Amount")
    ax2.set_xlabel("Category")
    ax2.set_ylabel(f"Amount ({currency_code})")
    plt.xticks(rotation=45, ha="right")
    st.pyplot(fig2)
    plt.close(fig2)

st.subheader("Auto Report Writing")
report_option = st.selectbox("Select Report Draft Type", options=list(report_sections.keys()))
st.markdown("### Generated Draft")
st.text_area("Report Text", value=report_sections[report_option], height=220)

full_report_text = "\n\n".join(f"{title}\n{content}" for title, content in report_sections.items())
st.download_button(
    label="📝 Download Report Draft (.txt)",
    data=full_report_text.encode("utf-8"),
    file_name="Project_Report_Draft.txt",
    mime="text/plain",
)

excel_file = create_excel_report(
    project_name=project_name,
    project_type=project_type,
    company_name=company_name,
    client_name=client_name,
    project_reference=project_reference,
    currency_code=currency_code,
    currency_symbol=currency_symbol,
    author_name=author_name,
    today=today,
    total_budget=total_budget,
    allocations=allocations,
    df=df,
    insights=insights,
    report_sections=report_sections,
    logo_path=logo_path_input.strip() or None,
)

pdf_file = create_pdf_report(
    project_name=project_name,
    project_type=project_type,
    company_name=company_name,
    client_name=client_name,
    project_reference=project_reference,
    currency_code=currency_code,
    currency_symbol=currency_symbol,
    author_name=author_name,
    today=today,
    total_budget=total_budget,
    allocations=allocations,
    df=df,
    insights=insights,
    report_sections=report_sections,
)

st.markdown("---")
st.markdown("### Report Metadata")
st.write(f"**Project Name:** {project_name}")
st.write(f"**Project Type:** {project_type}")
st.write(f"**Project Reference:** {project_reference}")
st.write(f"**Client Name:** {client_name}")
st.write(f"**Company Name:** {company_name}")
st.write(f"**Currency:** {currency_code}")
st.write(f"**Author:** {author_name}")
st.write(f"**Date:** {today}")
st.write(f"**Logo Path:** {logo_path_input if logo_path_input else 'Not provided'}")

export_col1, export_col2 = st.columns(2)
with export_col1:
    st.download_button(
        label="📥 Download Excel Report",
        data=excel_file,
        file_name="Project_Budget_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

with export_col2:
    st.download_button(
        label="📄 Download PDF Report",
        data=pdf_file,
        file_name="Project_Budget_Report.pdf",
        mime="application/pdf",
    )

st.success("Version 7 rebuilt cleanly with Excel export, PDF export, report writing, and branding support.")
