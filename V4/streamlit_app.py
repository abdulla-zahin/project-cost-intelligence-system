import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from datetime import date
from io import BytesIO

st.set_page_config(
    page_title="Project Cost Intelligence System",
    page_icon="🏗️",
    layout="wide",
)

# -----------------------------
# Project type recommendations
# -----------------------------
PROJECT_TYPE_SPLITS = {
    "General Contracting": {
        "Materials": 40,
        "Labor": 25,
        "Transportation": 5,
        "Office Expense": 5,
        "Salaries / Overheads": 10,
        "Company Profit": 10,
        "Contingency": 5,
    },
    "Pipeline Project": {
        "Materials": 50,
        "Labor": 20,
        "Transportation": 10,
        "Office Expense": 5,
        "Salaries / Overheads": 5,
        "Company Profit": 7,
        "Contingency": 3,
    },
    "Civil Construction": {
        "Materials": 45,
        "Labor": 30,
        "Transportation": 5,
        "Office Expense": 5,
        "Salaries / Overheads": 8,
        "Company Profit": 5,
        "Contingency": 2,
    },
    "Mechanical Installation": {
        "Materials": 38,
        "Labor": 32,
        "Transportation": 6,
        "Office Expense": 5,
        "Salaries / Overheads": 9,
        "Company Profit": 7,
        "Contingency": 3,
    },
    "Maintenance Contract": {
        "Materials": 25,
        "Labor": 40,
        "Transportation": 8,
        "Office Expense": 7,
        "Salaries / Overheads": 10,
        "Company Profit": 7,
        "Contingency": 3,
    },
    "EPC Project": {
        "Materials": 42,
        "Labor": 22,
        "Transportation": 6,
        "Office Expense": 6,
        "Salaries / Overheads": 12,
        "Company Profit": 8,
        "Contingency": 4,
    },
}


# -----------------------------
# Helper functions
# -----------------------------
def build_budget_dataframe(total_budget: float, allocations: dict) -> pd.DataFrame:
    rows = []
    for category, pct in allocations.items():
        amount = total_budget * (pct / 100)
        rows.append(
            {
                "Category": category,
                "Allocation %": round(pct, 2),
                "Amount": round(amount, 2),
            }
        )
    return pd.DataFrame(rows)


def analyze_budget(allocations: dict, total_pct: float) -> list[str]:
    insights = []

    if total_pct < 100:
        insights.append("⚠️ Total allocation is below 100%. Some budget is still unassigned.")
    elif total_pct > 100:
        insights.append("❌ Total allocation exceeds 100%. Adjust the percentages to match the budget.")
    else:
        insights.append("✅ Total allocation is perfectly balanced at 100%.")

    profit = allocations["Company Profit"]
    labor = allocations["Labor"]
    materials = allocations["Materials"]
    contingency = allocations["Contingency"]
    transport = allocations["Transportation"]

    if profit < 8:
        insights.append("⚠️ Profit margin is below 8%. This may be too tight for comfortable execution.")
    elif 8 <= profit <= 15:
        insights.append("✅ Profit margin looks healthy for a practical contracting estimate.")
    else:
        insights.append("⚠️ Profit margin is quite high. Double-check competitiveness and client expectations.")

    if labor < 20:
        insights.append("⚠️ Labor allocation is low. Execution quality or manpower coverage may suffer.")
    elif labor > 35:
        insights.append("⚠️ Labor allocation is very high. Verify productivity assumptions and manpower planning.")

    if materials < 30:
        insights.append("⚠️ Material allocation is low. Review BOQ realism and procurement assumptions.")
    elif materials > 60:
        insights.append("⚠️ Material allocation is very high. Check whether supply cost is dominating the project.")

    if contingency < 5:
        insights.append("⚠️ Contingency is low. This leaves little room for site surprises or change orders.")
    else:
        insights.append("✅ Contingency reserve provides a reasonable safety buffer.")

    if transport > 10:
        insights.append("⚠️ Transportation cost is high. Recheck logistics, fuel, and delivery assumptions.")

    return insights


def create_excel_report(
    project_name: str,
    project_type: str,
    company_name: str,
    author_name: str,
    today: str,
    total_budget: float,
    allocations: dict,
    df: pd.DataFrame,
    insights: list[str],
) -> BytesIO:
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference

    output = BytesIO()

    profit_pct = allocations["Company Profit"]
    profit_amount = total_budget * profit_pct / 100
    total_allocation_pct = round(sum(allocations.values()), 2)

    if total_allocation_pct == 100:
        allocation_status = "Balanced"
    elif total_allocation_pct < 100:
        allocation_status = "Under Allocated"
    else:
        allocation_status = "Over Allocated"

    insights_with_status = [f"Allocation Status: {allocation_status}"] + insights

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Write base data
        df.to_excel(writer, index=False, sheet_name="Budget Breakdown", startrow=5)
        pd.DataFrame({"Budget Insights": insights_with_status}).to_excel(
            writer, index=False, sheet_name="Budget Insights", startrow=3
        )

        wb = writer.book

        # -----------------------------
        # Theme / styles
        # -----------------------------
        dark_fill = PatternFill("solid", fgColor="1F4E78")
        mid_fill = PatternFill("solid", fgColor="D9EAF7")
        soft_fill = PatternFill("solid", fgColor="EDF4FB")
        gold_fill = PatternFill("solid", fgColor="FFF2CC")
        green_fill = PatternFill("solid", fgColor="E2F0D9")
        red_fill = PatternFill("solid", fgColor="FDE9E7")
        gray_fill = PatternFill("solid", fgColor="F3F3F3")

        white_bold = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(color="FFFFFF", bold=True, size=16)
        section_font = Font(bold=True, size=11)
        header_font = Font(bold=True)
        big_font = Font(bold=True, size=13)

        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

        thin = Side(style="thin", color="B7B7B7")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # -----------------------------
        # Executive Summary sheet
        # -----------------------------
        ws = wb.create_sheet("Executive Summary", 0)
        ws.sheet_view.showGridLines = False

        for col, width in {
            "A": 24,
            "B": 24,
            "C": 24,
            "D": 24,
            "E": 20,
            "F": 20,
        }.items():
            ws.column_dimensions[col].width = width

        ws.merge_cells("A1:F1")
        ws["A1"] = f"{company_name} - PROJECT COST INTELLIGENCE REPORT"
        ws["A1"].fill = dark_fill
        ws["A1"].font = title_font
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A3:C3")
        ws["A3"] = "PROJECT INFORMATION"
        ws["A3"].fill = mid_fill
        ws["A3"].font = section_font
        ws["A3"].alignment = center

        ws.merge_cells("D3:F3")
        ws["D3"] = "KEY FINANCIAL METRICS"
        ws["D3"].fill = gold_fill
        ws["D3"].font = section_font
        ws["D3"].alignment = center

        info_rows = [
            ("Project Name", project_name),
            ("Project Type", project_type),
            ("Company Name", company_name),
            ("Author", author_name),
            ("Date", today),
        ]

        metric_rows = [
            ("Total Budget", total_budget),
            ("Profit %", profit_pct / 100),
            ("Profit Amount", profit_amount),
            ("Allocation %", total_allocation_pct / 100),
            ("Status", allocation_status),
        ]

        start_row = 4

        for i, (label, value) in enumerate(info_rows, start=start_row):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            ws[f"A{i}"].fill = soft_fill
            ws[f"A{i}"].font = header_font
            ws[f"A{i}"].border = border
            ws[f"B{i}"].border = border
            ws[f"C{i}"].border = border
            ws[f"A{i}"].alignment = left
            ws[f"B{i}"].alignment = left

        for i, (label, value) in enumerate(metric_rows, start=start_row):
            ws[f"D{i}"] = label
            ws[f"E{i}"] = value
            ws[f"F{i}"].value = ""

            ws[f"D{i}"].fill = gray_fill
            ws[f"D{i}"].font = header_font

            ws[f"D{i}"].border = border
            ws[f"E{i}"].border = border
            ws[f"F{i}"].border = border

            ws[f"D{i}"].alignment = left
            ws[f"E{i}"].alignment = center if label != "Status" else left

            if label in {"Total Budget", "Profit Amount"}:
                ws[f"E{i}"].number_format = '#,##0.00'
                ws[f"E{i}"].font = big_font
            elif label in {"Profit %", "Allocation %"}:
                ws[f"E{i}"].number_format = '0.00%'
                ws[f"E{i}"].font = big_font
            elif label == "Status":
                ws[f"E{i}"].fill = green_fill if allocation_status == "Balanced" else red_fill
                ws[f"E{i}"].font = header_font

        ws.merge_cells("A11:F11")
        ws["A11"] = "EXECUTIVE NOTE"
        ws["A11"].fill = mid_fill
        ws["A11"].font = section_font
        ws["A11"].alignment = center

        note_text = (
            f"This report summarizes the recommended allocation for a {project_type.lower()} "
            f"under {company_name}. Current allocation status is {allocation_status.lower()}, "
            f"with an estimated profit of {profit_pct:.2f}% and profit amount of {profit_amount:,.2f}."
        )
        ws.merge_cells("A12:F13")
        ws["A12"] = note_text
        ws["A12"].alignment = wrap_left
        ws["A12"].border = border

        # -----------------------------
        # Budget Breakdown sheet
        # -----------------------------
        bd = wb["Budget Breakdown"]
        bd.sheet_view.showGridLines = False

        for col, width in {"A": 28, "B": 16, "C": 20}.items():
            bd.column_dimensions[col].width = width

        bd.merge_cells("A1:C1")
        bd["A1"] = "DETAILED BUDGET BREAKDOWN"
        bd["A1"].fill = dark_fill
        bd["A1"].font = title_font
        bd["A1"].alignment = center

        bd["A3"] = "Project Name"
        bd["B3"] = project_name
        bd["A4"] = "Project Type"
        bd["B4"] = project_type

        for cell in ["A3", "A4"]:
            bd[cell].fill = soft_fill
            bd[cell].font = header_font

        for cell in ["A3", "B3", "A4", "B4"]:
            bd[cell].border = border

        table_header_row = 6
        for cell in bd[table_header_row]:
            cell.fill = mid_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for row in range(table_header_row + 1, table_header_row + 1 + len(df)):
            bd[f"B{row}"].number_format = '0.00'
            bd[f"C{row}"].number_format = '#,##0.00'

            for col in "ABC":
                bd[f"{col}{row}"].border = border

            category = bd[f"A{row}"].value

            if category == "Company Profit":
                for col in "ABC":
                    bd[f"{col}{row}"].fill = gold_fill
                    bd[f"{col}{row}"].font = header_font

            elif category == "Contingency":
                for col in "ABC":
                    bd[f"{col}{row}"].fill = green_fill

        total_row = table_header_row + 1 + len(df)
        bd[f"A{total_row}"] = "Total Allocation"
        bd[f"B{total_row}"] = total_allocation_pct
        bd[f"C{total_row}"] = df["Amount"].sum()

        for col in "ABC":
            bd[f"{col}{total_row}"].fill = gray_fill
            bd[f"{col}{total_row}"].font = header_font
            bd[f"{col}{total_row}"].border = border

        bd[f"B{total_row}"].number_format = '0.00'
        bd[f"C{total_row}"].number_format = '#,##0.00'

        bd.freeze_panes = "A6"

        # -----------------------------
        # Budget Insights sheet
        # -----------------------------
        ins = wb["Budget Insights"]
        ins.sheet_view.showGridLines = False
        ins.column_dimensions["A"].width = 110

        ins["A1"] = "BUDGET INTELLIGENCE & RISK NOTES"
        ins["A1"].fill = dark_fill
        ins["A1"].font = title_font
        ins["A1"].alignment = center

        ins["A4"].fill = mid_fill
        ins["A4"].font = header_font
        ins["A4"].border = border

        for row in range(5, 5 + len(insights_with_status)):
            ins[f"A{row}"].border = border
            ins[f"A{row}"].alignment = wrap_left
            value = str(ins[f"A{row}"].value)

            if "✅" in value or "Balanced" in value:
                ins[f"A{row}"].fill = green_fill
            elif "⚠️" in value or "❌" in value or "Over" in value or "Under" in value:
                ins[f"A{row}"].fill = red_fill

        ins.freeze_panes = "A4"

        # -----------------------------
        # Charts sheet
        # -----------------------------
        ch = wb.create_sheet("Charts")
        ch.sheet_view.showGridLines = False

        ch.merge_cells("A1:H1")
        ch["A1"] = "BUDGET VISUALS"
        ch["A1"].fill = dark_fill
        ch["A1"].font = title_font
        ch["A1"].alignment = center

        ch["A2"] = "Category"
        ch["B2"] = "Amount"
        ch["C2"] = "Allocation %"

        for cell in ch[2]:
            cell.fill = mid_fill
            cell.font = header_font
            cell.border = border

        for idx, (_, row) in enumerate(df.iterrows(), start=3):
            ch[f"A{idx}"] = row["Category"]
            ch[f"B{idx}"] = row["Amount"]
            ch[f"C{idx}"] = row["Allocation %"] / 100

        # Pie Chart
        pie = PieChart()
        pie_labels = Reference(ch, min_col=1, min_row=3, max_row=2 + len(df))
        pie_data = Reference(ch, min_col=2, min_row=2, max_row=2 + len(df))

        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(pie_labels)
        pie.title = "Budget Distribution"
        pie.height = 9
        pie.width = 12
        ch.add_chart(pie, "E3")

        # Bar Chart
        bar = BarChart()
        bar.type = "col"
        bar.style = 10
        bar.title = "Allocation by Category"
        bar.y_axis.title = "Amount"
        bar.x_axis.title = "Category"

        bar_data = Reference(ch, min_col=2, min_row=2, max_row=2 + len(df))
        bar_labels = Reference(ch, min_col=1, min_row=3, max_row=2 + len(df))

        bar.add_data(bar_data, titles_from_data=True)
        bar.set_categories(bar_labels)
        bar.height = 9
        bar.width = 14
        ch.add_chart(bar, "E20")

    output.seek(0)
    return output


# -----------------------------
# Header
# -----------------------------
st.title("🏗️ Project Cost Intelligence System")
st.caption(
    "Smart budget allocation, profit analysis, and planning dashboard for engineering / contracting projects."
)

# -----------------------------
# Project details
# -----------------------------
st.subheader("Project Details")
col1, col2 = st.columns(2)

with col1:
    project_name = st.text_input("Project Name", value="New Engineering Project")
    company_name = st.text_input("Company Name", value="TKM")

with col2:
    author_name = "Abdulla Zahin"
    today = date.today().strftime("%d %B %Y")
    st.text_input("Author", value=author_name, disabled=True)
    st.text_input("Date", value=today, disabled=True)

# -----------------------------
# Budget input
# -----------------------------
st.subheader("Budget Input")

project_type = st.selectbox(
    "Project Type",
    options=list(PROJECT_TYPE_SPLITS.keys()),
    index=0,
)

if "last_project_type" not in st.session_state:
    st.session_state.last_project_type = project_type

if "allocations_state" not in st.session_state:
    st.session_state.allocations_state = PROJECT_TYPE_SPLITS[project_type].copy()

if project_type != st.session_state.last_project_type:
    st.session_state.allocations_state = PROJECT_TYPE_SPLITS[project_type].copy()
    st.session_state.last_project_type = project_type

selected_split = st.session_state.allocations_state

total_budget = st.number_input(
    "Total Project Budget",
    min_value=0.0,
    value=500000.0,
    step=10000.0,
    format="%.2f",
)

st.markdown("### Recommended Allocation")
st.write(f"Smart recommendation is loaded for **{project_type}**. You can still adjust the percentages if needed.")

alloc_col1, alloc_col2 = st.columns(2)

allocations = {}
items = list(selected_split.items())
mid = len(items) // 2 + len(items) % 2

with alloc_col1:
    for category, value in items[:mid]:
        widget_key = f"alloc_{category}"
        if widget_key not in st.session_state:
            st.session_state[widget_key] = float(value)

        allocations[category] = st.number_input(
            f"{category} (%)",
            min_value=0.0,
            max_value=100.0,
            value=float(st.session_state.allocations_state[category]),
            step=1.0,
            key=widget_key,
        )
        st.session_state.allocations_state[category] = allocations[category]

with alloc_col2:
    for category, value in items[mid:]:
        widget_key = f"alloc_{category}"
        if widget_key not in st.session_state:
            st.session_state[widget_key] = float(value)

        allocations[category] = st.number_input(
            f"{category} (%)",
            min_value=0.0,
            max_value=100.0,
            value=float(st.session_state.allocations_state[category]),
            step=1.0,
            key=widget_key,
        )
        st.session_state.allocations_state[category] = allocations[category]

total_pct = sum(allocations.values())
remaining_pct = 100 - total_pct

info1, info2, info3, info4 = st.columns(4)
info1.metric("Total Allocation %", f"{total_pct:.2f}%")
info2.metric("Unallocated / Excess %", f"{remaining_pct:.2f}%")
info3.metric("Estimated Profit %", f"{allocations['Company Profit']:.2f}%")
info4.metric("Estimated Profit Amount", f"{total_budget * allocations['Company Profit'] / 100:,.2f}")

# -----------------------------
# Data table
# -----------------------------
df = build_budget_dataframe(total_budget, allocations)

st.subheader("Budget Breakdown")
st.dataframe(df, use_container_width=True, hide_index=True)

# -----------------------------
# Analysis
# -----------------------------
st.subheader("Budget Intelligence")
insights = analyze_budget(allocations, total_pct)
for item in insights:
    st.write(item)

# -----------------------------
# Summary cards
# -----------------------------
st.subheader("Project Summary")
summary1, summary2, summary3, summary4 = st.columns(4)

profit_amount = total_budget * allocations["Company Profit"] / 100
materials_amount = total_budget * allocations["Materials"] / 100
labor_amount = total_budget * allocations["Labor"] / 100
contingency_amount = total_budget * allocations["Contingency"] / 100

summary1.metric("Total Budget", f"{total_budget:,.2f}")
summary2.metric("Materials", f"{materials_amount:,.2f}")
summary3.metric("Labor", f"{labor_amount:,.2f}")
summary4.metric("Contingency", f"{contingency_amount:,.2f}")

st.metric("Company Profit", f"{profit_amount:,.2f}", f"{allocations['Company Profit']:.2f}% margin")

# -----------------------------
# Charts in Streamlit
# -----------------------------
st.subheader("Visual Dashboard")
chart_col1, chart_col2 = st.columns(2)

with chart_col1:
    fig1, ax1 = plt.subplots(figsize=(6, 6))
    ax1.pie(df["Amount"], labels=df["Category"], autopct="%1.1f%%", startangle=90)
    ax1.set_title("Budget Distribution")
    st.pyplot(fig1)

with chart_col2:
    fig2, ax2 = plt.subplots(figsize=(8, 5))
    ax2.bar(df["Category"], df["Amount"])
    ax2.set_title("Category-wise Budget Amount")
    ax2.set_xlabel("Category")
    ax2.set_ylabel("Amount")
    plt.xticks(rotation=45, ha="right")
    st.pyplot(fig2)

# -----------------------------
# Excel report
# -----------------------------
excel_file = create_excel_report(
    project_name=project_name,
    project_type=project_type,
    company_name=company_name,
    author_name=author_name,
    today=today,
    total_budget=total_budget,
    allocations=allocations,
    df=df,
    insights=insights,
)

# -----------------------------
# Report footer
# -----------------------------
st.markdown("---")
st.markdown("### Report Metadata")
st.write(f"**Project Name:** {project_name}")
st.write(f"**Project Type:** {project_type}")
st.write(f"**Company Name:** {company_name}")
st.write(f"**Author:** {author_name}")
st.write(f"**Date:** {today}")

st.download_button(
    label="📥 Download Excel Report",
    data=excel_file,
    file_name="Project_Budget_Report.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.success("Project Cost Intelligence System is ready for estimation, analysis, and Excel reporting.")